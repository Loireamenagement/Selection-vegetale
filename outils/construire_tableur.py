# -*- coding: utf-8 -*-
import pandas as pd, urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fiches import FICHES
from fiches2 import FICHES2, ALIAS as ALIAS2
from fiches3 import FICHES3, ALIAS3
from fiches4 import FICHES4, ALIAS4
from fiches5 import FICHES5, ALIAS5

ALIAS = {**ALIAS2, **ALIAS3, **ALIAS4, **ALIAS5}
from complements import COMPLEMENTS

FICHES_TOUTES = []
for _f in FICHES:
    _d = dict(_f)
    _b, _dn, _e = COMPLEMENTS[_f["lat"]]
    _d["bdm"], _d["dens"], _d["esp"] = _b, _dn, _e
    FICHES_TOUTES.append(_d)
FICHES_TOUTES.extend(dict(_f) for _f in FICHES2)
FICHES_TOUTES.extend(dict(_f) for _f in FICHES3)
FICHES_TOUTES.extend(dict(_f) for _f in FICHES4)
FICHES_TOUTES.extend(dict(_f) for _f in FICHES5)

cat = pd.read_pickle("/home/claude/cat.pkl")
fmap = {f["lat"]: f for f in FICHES_TOUTES}
for _variante, _cible in ALIAS.items():          # variantes d'ecriture du catalogue
    if _cible in fmap:
        fmap[_variante] = fmap[_cible]

MOIS = ["", "janvier", "février", "mars", "avril", "mai", "juin", "juillet",
        "août", "septembre", "octobre", "novembre", "décembre"]

g = cat.groupby("Nom latin", as_index=False).agg(
    **{"Nom commun (FR)": ("Nom commun (FR)", "first"),
       "Sections": ("Section", lambda s: " / ".join(sorted(set(s)))),
       "Nb références": ("Nom complet", "size"),
       "Nb cultivars": ("Cultivar", lambda s: len(set(x for x in s if x))),
       "Conteneurs dispo": ("Conteneur", lambda s: ", ".join(sorted(set(x for x in s if x)))),
       "Prix HT min": ("Prix HT palier 1", "min"),
       "Prix HT max": ("Prix HT palier 1", "max")})

lignes = []
for _, r in g.iterrows():
    lat = r["Nom latin"]
    f = fmap.get(lat)
    q = urllib.parse.quote(lat)
    d = {
        "Fiche complétée": "Oui" if f else "",
        "Nom latin": lat,
        "Nom commun (FR)": f["fr"] if f else r["Nom commun (FR)"],
        "Type": f["type"] if f else "",
        "Hauteur adulte (m)": f["h"] if f else "",
        "Largeur adulte (m)": f["l"] if f else "",
        "Exposition": f["expo"] if f else "",
        "Nature de sol": f["sol"] if f else "",
        "Humidité du sol": f["hum"] if f else "",
        "Rusticité (°C)": f["rust"] if f else "",
        "Bord de mer": f["bdm"] if f else "",
        "Feuillage": f["feu"] if f else "",
        "Couleur feuillage": f["cfeu"] if f else "",
        "Floraison début (mois)": f["m1"] if f and f["m1"] else "",
        "Floraison fin (mois)": f["m2"] if f and f["m2"] else "",
        "Période de floraison": f["flo"] if f else "",
        "Couleur floraison": f["cflo"] if f else "",
        "Période de plantation": f["plant"] if f else "",
        "Période optimale": f["popt"] if f else "",
        "Niveau d'entretien": f["ent"] if f else "",
        "Densité (sujets/m²)": f["dens"] if f else "",
        "Distance de plantation (m)": f["esp"] if f else "",
        "Style / usages": f["style"] if f else "",
        "Associations": f["asso"] if f else "",
        "Conseils d'entretien": f["cons"] if f else "",
        "Recherche photo (Wikimedia)": f"https://commons.wikimedia.org/w/index.php?search={q}&title=Special:MediaSearch&type=image",
        "Photo retenue (URL)": "",
        "Notes perso": "",
        "Sections catalogue": r["Sections"],
        "Nb références": r["Nb références"],
        "Nb cultivars": r["Nb cultivars"],
        "Conteneurs dispo": r["Conteneurs dispo"],
        "Prix HT min": r["Prix HT min"],
        "Prix HT max": r["Prix HT max"],
    }
    lignes.append(d)

especes = pd.DataFrame(lignes)
especes["_ord"] = especes["Fiche complétée"].eq("Oui").map({True: 0, False: 1})
especes = especes.sort_values(["_ord", "Nom latin"]).drop(columns="_ord").reset_index(drop=True)

COLS_CATALOGUE = ["Section", "Nom latin", "Cultivar", "Nom complet", "Nom commun (FR)",
                  "Genre", "Conteneur", "Taille", "Dispo", "Prix HT palier 1",
                  "Prix HT palier 2", "Prix HT palier 3", "Origine"]
catalogue = cat[COLS_CATALOGUE].copy()

# --- Onglet referentiel des valeurs de filtre ---
REFERENTIEL = [
    ("Type", "Arbre · Arbuste · Conifère · Couvre-sol · Vivace · Graminée · Grimpante · Fougère · Fruitier · Aromatique"),
    ("Exposition", "Plein soleil · Soleil à mi-ombre · Mi-ombre · Mi-ombre à ombre · Soleil à ombre · Ombre"),
    ("Nature de sol", "Tout type · Drainé · Profond · Humifère · Sec · Frais · Calcaire toléré · Acide (terre de bruyère) · Argileux toléré · Pauvre toléré · Caillouteux"),
    ("Humidité du sol", "Sec · Sec à frais · Frais · Frais à humide · Humide"),
    ("Rusticité (°C)", "Valeur numérique négative : -30 · -25 · -20 · -18 · -15 · -12 · -10"),
    ("Bord de mer", "Front de mer (supporte les embruns directs) · Second rideau (derrière une haie brise-vent) · Déconseillé en bord de mer"),
    ("Feuillage", "Persistant · Semi-persistant · Caduc"),
    ("Couleur feuillage", "Vert · Vert foncé · Vert clair · Vert grisé · Vert bleuté · Argenté · Pourpre · Doré · Panaché · Rouge automnal"),
    ("Couleur floraison", "Blanc · Crème · Rose · Rouge · Corail · Orange · Jaune · Mauve · Violet · Bleu · Vert · Bicolore"),
    ("Floraison début / fin", "Numéro de mois de 1 à 12. Permet le filtre « qu'est-ce qui fleurit en août ? »"),
    ("Période de plantation", "Automne · Printemps · Automne à printemps · Printemps (à éviter en automne)"),
    ("Niveau d'entretien", "Faible · Moyen · Soutenu"),
    ("Densité (sujets/m²)", "Valeur numérique. Plantation en masse : 4 à 9 pour un couvre-sol, 3 à 6 pour une vivace, 1 à 3 pour une graminée, 0,3 à 1 pour un arbuste, moins de 0,1 pour un arbre."),
    ("Distance de plantation (m)", "Valeur numérique. Distance entre deux sujets. Sert au calcul en ligne (haie, alignement) là où la densité sert au calcul en surface."),
]

# --- Ecriture ---
wb = Workbook()
FONT = "Arial"
VERT = PatternFill("solid", fgColor="1F5C3A")
VERT_CLAIR = PatternFill("solid", fgColor="E8F1EC")
JAUNE = PatternFill("solid", fgColor="FFF2CC")
GRIS = PatternFill("solid", fgColor="F2F2F2")
BORD = Border(*[Side(style="thin", color="D0D0D0")] * 4)

def ecrire(ws, df, cols_jaunes=()):
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = VERT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for _, row in df.iterrows():
        ws.append(list(row.values))
    idx_jaunes = {list(df.columns).index(c) + 1 for c in cols_jaunes if c in df.columns}
    for r in range(2, ws.max_row + 1):
        rempli = ws.cell(row=r, column=1).value == "Oui"
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORD
            cell.alignment = Alignment(vertical="top", wrap_text=(c in idx_jaunes))
            if c in idx_jaunes and not rempli:
                cell.fill = JAUNE
            elif c in idx_jaunes and rempli:
                cell.fill = VERT_CLAIR
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 34
    for c, col in enumerate(df.columns, start=1):
        ech = [len(str(v)) for v in df[col].head(200)]
        largeur = max([len(str(col))] + ech) + 2
        ws.column_dimensions[get_column_letter(c)].width = min(max(largeur, 11), 46)

COLS_HORTI = ["Type", "Hauteur adulte (m)", "Largeur adulte (m)", "Exposition",
              "Nature de sol", "Humidité du sol", "Rusticité (°C)", "Bord de mer", "Feuillage",
              "Couleur feuillage", "Floraison début (mois)", "Floraison fin (mois)",
              "Période de floraison", "Couleur floraison", "Période de plantation",
              "Période optimale", "Niveau d'entretien", "Densité (sujets/m²)",
              "Distance de plantation (m)", "Style / usages",
              "Associations", "Conseils d'entretien", "Photo retenue (URL)", "Notes perso"]

ws1 = wb.active
ws1.title = "Fiches espèces"
ecrire(ws1, especes, cols_jaunes=COLS_HORTI)
ws1.auto_filter.ref = f"A1:{get_column_letter(len(especes.columns))}{ws1.max_row}"
cols = list(especes.columns)
for r in range(2, ws1.max_row + 1):
    for nom in ("Prix HT min", "Prix HT max"):
        ws1.cell(row=r, column=cols.index(nom) + 1).number_format = '#,##0.00 €'

ws2 = wb.create_sheet("Catalogue fournisseur")
ecrire(ws2, catalogue)
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(catalogue.columns))}{ws2.max_row}"
for r in range(2, ws2.max_row + 1):
    for c in (10, 11, 12):
        ws2.cell(row=r, column=c).number_format = '#,##0.00 €'

ws3 = wb.create_sheet("Référentiel filtres")
ws3.append(["Champ", "Valeurs autorisées"])
for c in (1, 2):
    cell = ws3.cell(row=1, column=c)
    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    cell.fill = VERT
ws3.column_dimensions["A"].width = 26
ws3.column_dimensions["B"].width = 110
for champ, vals in REFERENTIEL:
    ws3.append([champ, vals])
for r in range(2, ws3.max_row + 1):
    ws3.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=True)
    ws3.cell(row=r, column=2).font = Font(name=FONT, size=10)
    ws3.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

ws4 = wb.create_sheet("Notice")
notice = [
    ("BASE VÉGÉTALE — LOT 1", ""),
    ("Source catalogue", "AD.V Production – Paysage 2026/2027"),
    ("Date", "27/08/2026"),
    ("", ""),
    ("ÉTAT D'AVANCEMENT", ""),
    ("Espèces au total", f"{len(especes)} espèces botaniques issues du catalogue"),
    ("Fiches complétées", f"{len(FICHES_TOUTES)} — les espèces les plus représentées chez AD.V (colonne « Fiche complétée » = Oui, fond vert)"),
    ("Fiches à compléter", f"{len(especes) - len(FICHES_TOUTES)} — cellules sur fond jaune"),
    ("", ""),
    ("DEUX CHAMPS DE CHIFFRAGE", ""),
    ("Bord de mer", "Trois niveaux. « Front de mer » supporte les embruns directs, « Second rideau » demande la protection d'une haie brise-vent, « Déconseillé » souffre du sel et du vent. Critère déterminant sur le littoral atlantique."),
    ("Densité et distance", "La densité (sujets/m²) sert à chiffrer une surface : massif, couvre-sol, talus. La distance de plantation sert à chiffrer une longueur : haie, alignement, bordure. Les deux sont nécessaires, aucune ne remplace l'autre."),
    ("", ""),
    ("ORGANISATION DU FICHIER", ""),
    ("Fiches espèces", "Une ligne par espèce botanique. C'est la source des fiches techniques de l'application. Les 50 espèces traitées sont remontées en tête de liste."),
    ("Catalogue fournisseur", f"{len(catalogue)} références commerciales : une ligne par combinaison espèce / cultivar / conteneur / taille, avec les trois paliers de prix."),
    ("Référentiel filtres", "Les valeurs autorisées pour chaque champ. À respecter impérativement : ce sont elles qui deviendront les filtres de l'application. Une valeur saisie hors référentiel ne sera pas filtrable."),
    ("", ""),
    ("ORIGINE DES DONNÉES HORTICOLES", ""),
    ("Nature", "Les données horticoles ne proviennent pas du catalogue fournisseur, qui ne contient que des informations commerciales. Elles ont été établies à partir de connaissances horticoles générales."),
    ("Validation", "Elles constituent un premier jet professionnel à valider avant diffusion à un client. La rusticité est la donnée la plus variable selon les sources : les valeurs indiquées sont prudentes et supposent un sol correctement drainé."),
    ("Cultivars", "Les fiches décrivent l'espèce type. Pour les cultivars nains, compacts ou panachés, ajuster la taille adulte et vérifier la rusticité, souvent inférieure."),
    ("", ""),
    ("PHOTOS", ""),
    ("Colonne Wikimedia", "Chaque espèce dispose d'un lien de recherche vers Wikimedia Commons. Ces images sont sous licence libre et utilisables commercialement, à condition de créditer l'auteur."),
    ("Recommandation", "Demander à AD.V leur photothèque : ce sont les photos des variétés réellement produites, et les pépiniéristes les cèdent en général volontiers à leurs clients revendeurs."),
    ("À éviter", "Les images issues de Google Images ou de catalogues concurrents sont protégées par le droit d'auteur et ne peuvent pas être utilisées dans un outil commercial."),
    ("", ""),
    ("POINTS DE VIGILANCE", ""),
    ("Noms latins", "Le découpage espèce / cultivar est automatique. Sur les sections rédigées en majuscules dans le catalogue d'origine, quelques cultivars ont pu être classés comme espèces (ex. « Abelia edward »). Une relecture est recommandée, cette colonne servant de clé de liaison avec une base botanique externe."),
    ("Prix", "Prix de la saison 2026/2027, à actualiser chaque année. Les seuils de quantité des trois paliers varient selon les sections du catalogue d'origine."),
    ("Disponibilités", "La colonne « Dispo » n'est renseignée que sur une partie du catalogue d'origine."),
]
ws4.column_dimensions["A"].width = 26
ws4.column_dimensions["B"].width = 108
for i, (a, b) in enumerate(notice, start=1):
    ca = ws4.cell(row=i, column=1, value=a)
    ca.font = Font(name=FONT, bold=(b == ""), size=11 if b == "" else 10)
    ca.alignment = Alignment(vertical="top")
    cb = ws4.cell(row=i, column=2, value=b)
    cb.font = Font(name=FONT, size=10)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if b == "" and a:
        ca.fill = GRIS

wb.save("/mnt/user-data/outputs/Base_vegetale_AD-V.xlsx")
print("OK -", len(especes), "especes dont", sum(1 for l in lignes if l["Fiche complétée"]=="Oui"), "completees")
